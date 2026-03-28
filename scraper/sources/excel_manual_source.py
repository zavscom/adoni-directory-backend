"""Load manually edited rows from data/adoni_directory.xlsx (Manual Additions sheet)."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scraper import config
from scraper.excel_headers import canonical_excel_header, is_known_excel_field
from scraper.sources.base_source import BaseSource

logger = logging.getLogger(__name__)

EXCEL_FILENAME = "adoni_directory.xlsx"
MANUAL_SHEET = "Manual Additions"


class ExcelManualSource(BaseSource):
    """Local Excel rows; registered after web sources so shared [id] can merge in dedupe."""

    @property
    def source_key(self) -> str:
        return "excel_manual"

    def fetch_raw(self) -> list[dict[str, Any]]:
        path: Path = config.DATA_DIR / EXCEL_FILENAME
        if not path.is_file():
            logger.info("excel_manual: no file at %s (skipping)", path)
            return []

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            logger.warning("excel_manual: could not open %s: %s (skipping)", path, e)
            return []

        try:
            if MANUAL_SHEET not in wb.sheetnames:
                logger.warning("excel_manual: sheet %r missing in %s", MANUAL_SHEET, path)
                return []
            ws = wb[MANUAL_SHEET]
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return []

            header_keys: list[str | None] = [canonical_excel_header(c) for c in header_row]

            out: list[dict[str, Any]] = []
            for row in rows_iter:
                if row is None:
                    continue
                record: dict[str, Any] = {}
                for idx, key in enumerate(header_keys):
                    if key is None or not is_known_excel_field(key):
                        continue
                    if idx >= len(row):
                        continue
                    val = row[idx]
                    if val is None:
                        continue
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    record[key] = val

                name = str(record.get("name", "")).strip()
                if not name:
                    continue

                if "extra_json" in record:
                    raw_ex = record.pop("extra_json")
                    if isinstance(raw_ex, str) and raw_ex.strip():
                        try:
                            record["extra"] = json.loads(raw_ex)
                        except json.JSONDecodeError:
                            record["extra"] = {}
                    elif isinstance(raw_ex, dict):
                        record["extra"] = raw_ex

                out.append(record)

            logger.info("excel_manual: loaded %d rows from %s", len(out), path)
            return out
        finally:
            wb.close()
