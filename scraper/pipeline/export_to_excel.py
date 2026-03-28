"""Export merged directory to Excel for local editing (Manual Additions sheet)."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from scraper.excel_headers import EXCEL_BUSINESS_HEADERS
from scraper.models import Business

logger = logging.getLogger(__name__)

EXPORT_HEADERS: list[str] = EXCEL_BUSINESS_HEADERS

DEFAULT_CATEGORY_SUGGESTIONS: tuple[str, ...] = (
    "Hospital",
    "Hospitals",
    "Clinic",
    "Doctor",
    "School",
    "Shop",
    "Restaurant",
    "Bank",
    "Other",
)

_HEADER_FONT = Font(bold=True)
_MAX_WIDTH = 55


def _is_excel_manual_business(b: Business) -> bool:
    parts = {s.strip() for s in b.source.split(",") if s.strip()}
    return "excel_manual" in parts


def _business_to_row(b: Business) -> list[Any]:
    extra_json = json.dumps(b.extra, ensure_ascii=False) if b.extra else ""
    return [
        b.id,
        b.name,
        b.category,
        b.subCategory or "",
        b.address,
        b.area,
        b.pincode,
        b.phone or "",
        b.whatsapp or "",
        b.email or "",
        b.website or "",
        b.city,
        b.state,
        b.latitude if b.latitude is not None else "",
        b.longitude if b.longitude is not None else "",
        b.source,
        b.lastSeenAt,
        extra_json,
    ]


def _autosize_columns(ws: Any, min_w: float = 10.0) -> None:
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        letter = get_column_letter(col_idx)
        max_len = min_w
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, min(len(str(cell.value)), _MAX_WIDTH))
        ws.column_dimensions[letter].width = max_len + 2


def _style_header_row(ws: Any, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = _HEADER_FONT


def _category_list_for_validation(businesses: Sequence[Business]) -> list[str]:
    seen: dict[str, None] = {}
    for b in businesses:
        c = (b.category or "").strip()
        if c:
            seen[c] = None
    for d in DEFAULT_CATEGORY_SUGGESTIONS:
        seen.setdefault(d, None)
    return sorted(seen.keys(), key=str.casefold)


def _apply_category_dropdown(
    ws: Any,
    categories: Sequence[str],
    header_row: int,
    category_col: int,
    data_start_row: int,
    max_row: int,
) -> None:
    if not categories or max_row < data_start_row:
        return
    joined = ",".join(c.replace('"', '""') for c in categories)
    if len(joined) > 240:
        joined = ",".join(c.replace('"', '""') for c in categories[:40]) + ",Other"
    formula = f'"{joined}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Pick a category from the list or leave blank."
    dv.errorTitle = "Invalid category"
    col_letter = get_column_letter(category_col)
    dv.add(f"{col_letter}{data_start_row}:{col_letter}{max_row}")
    ws.add_data_validation(dv)


def export_businesses_to_excel(
    businesses: list[Business],
    output_path: str | Path = "data/adoni_directory.xlsx",
) -> None:
    """
    Write adoni_directory.xlsx with:
    - All Data: full merged directory (scraped + manual).
    - Manual Additions: only businesses whose source includes excel_manual (else header only).
    - Template: headers only (blank template for new rows).
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    manual_businesses = [b for b in businesses if _is_excel_manual_business(b)]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Data"
    ws_all.append(EXPORT_HEADERS)
    _style_header_row(ws_all)
    for b in businesses:
        ws_all.append(_business_to_row(b))
    _autosize_columns(ws_all)

    ws_manual = wb.create_sheet("Manual Additions")
    ws_manual.append(EXPORT_HEADERS)
    _style_header_row(ws_manual)
    for b in manual_businesses:
        ws_manual.append(_business_to_row(b))
    _autosize_columns(ws_manual)
    cat_col = EXPORT_HEADERS.index("category") + 1
    max_m = max(ws_manual.max_row, 2)
    _apply_category_dropdown(
        ws_manual,
        _category_list_for_validation(businesses),
        header_row=1,
        category_col=cat_col,
        data_start_row=2,
        max_row=max_m,
    )

    ws_tpl = wb.create_sheet("Template")
    ws_tpl.append(EXPORT_HEADERS)
    _style_header_row(ws_tpl)
    _autosize_columns(ws_tpl)
    _apply_category_dropdown(
        ws_tpl,
        _category_list_for_validation(businesses),
        header_row=1,
        category_col=cat_col,
        data_start_row=2,
        max_row=502,
    )

    wb.save(path)
    logger.info(
        "export_excel: wrote %s (all=%d, manual_sheet=%d)",
        path,
        len(businesses),
        len(manual_businesses),
    )
